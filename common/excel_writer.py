"""
Excel 输出工具 — 多 Sheet 工作簿生成 + 图表插入。
"""

import os
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

from common import setup_logging

logger = setup_logging("excel_writer")


def auto_column_width(ws, min_width: int = 8, max_width: int = 50):
    """自动调整列宽。"""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        lengths = []
        for cell in col_cells:
            if cell.value:
                # 中文字符约等于 2 个英文字符宽度
                val = str(cell.value)
                length = sum(2 if ord(c) > 127 else 1 for c in val)
                lengths.append(length)
        if lengths:
            best = min(max(max(lengths) + 2, min_width), max_width)
            ws.column_dimensions[col_letter].width = best


class ExcelReport:
    """
    多 Sheet Excel 报告构建器。

    用法:
        report = ExcelReport("output.xlsx")
        report.add_sheet("原始数据", df_raw)
        report.add_sheet("统计指标", df_stats)
        report.add_sheet("频次统计", df_freq, chart_path="chart.png")
        report.save()
    """

    def __init__(self, filepath: str | Path):
        self.filepath = Path(filepath)
        self.filepath.parent.mkdir(parents=True, exist_ok=True)
        self.sheets: list[dict] = []  # {name, df, chart_path, auto_width}

    def add_sheet(self, name: str, df: pd.DataFrame,
                  chart_path: str | Path | None = None,
                  chart_position: str = "E1",
                  chart_width: int = 950, chart_height: int = 500,
                  auto_width: bool = True):
        """添加一个 Sheet。chart_path 非空则在指定位置插入图片。"""
        self.sheets.append({
            "name": name,
            "df": df,
            "chart_path": str(chart_path) if chart_path else None,
            "chart_position": chart_position,
            "chart_width": chart_width,
            "chart_height": chart_height,
            "auto_width": auto_width,
        })

    def save(self):
        """写入 Excel 文件并插入图表。"""
        # 第一步：写入数据
        with pd.ExcelWriter(self.filepath, engine="openpyxl") as writer:
            for sheet in self.sheets:
                sheet["df"].to_excel(writer, sheet_name=sheet["name"], index=False)

        # 第二步：插入图表 + 调整列宽
        wb = load_workbook(self.filepath)
        for sheet in self.sheets:
            ws = wb[sheet["name"]]
            if sheet["auto_width"]:
                auto_column_width(ws)

            chart_path = sheet.get("chart_path")
            if chart_path and os.path.exists(chart_path):
                img = Image(chart_path)
                img.width = sheet.get("chart_width", 950)
                img.height = sheet.get("chart_height", 500)
                ws.add_image(img, sheet.get("chart_position", "E1"))

        wb.save(self.filepath)
        wb.close()
        logger.info("Excel 已保存: %s", self.filepath)

    def cleanup_charts(self):
        """删除所有关联的临时图表文件。"""
        for sheet in self.sheets:
            path = sheet.get("chart_path")
            if path and os.path.exists(path):
                os.remove(path)


def build_stats_sheet(stats: dict, code: str, name: str,
                      years: int = 3, multi_period: bool = False) -> pd.DataFrame:
    """
    生成「统计指标与标准差」Sheet 的 DataFrame。

    参数
    ----
    multi_period : 是否展示 1/2/3 年多周期指标
    """
    rows: list[list] = []

    def add_row(indicator: str, value: str, note: str = "-"):
        rows.append([indicator, value, note])

    label = "ETF" if code.startswith("15") or code.startswith("51") or code.startswith("58") else "股票"

    add_row(f"{label}代码", code)
    add_row(f"{label}名称", name)
    add_row(f"价格均值（近{years}年）", f"{stats['mean']:.2f} 元")
    add_row(f"价格中位数（近{years}年）", f"{stats['median']:.2f} 元")
    add_row(f"价格众数（近{years}年）", f"{stats['mode']:.2f} 元", "出现频次最高的收盘价")
    add_row(f"价格标准差（近{years}年）", f"{stats['std']:.2f} 元")
    add_row(f"数据总交易日数（近{years}年）", f"{stats['count']} 天")

    if multi_period:
        for period, label_p in [("1y", "近1年"), ("2y", "近2年"), ("3y", "近3年")]:
            min_key = f"price_min_{period}"
            max_key = f"price_max_{period}"
            crit_key = f"critical_buy_point_{period}"
            prem_key = f"buy_point_premium_{period}"
            min_date_key = f"min_date_str_{period}"
            max_date_key = f"max_date_str_{period}"

            if min_key in stats:
                add_row(f"价格最小值（{label_p}）", f"{stats[min_key]:.2f} 元",
                        f"{label_p}最低收盘价, 日期: {stats.get(min_date_key, '未知')}")
            if max_key in stats:
                add_row(f"价格最大值（{label_p}）", f"{stats[max_key]:.2f} 元",
                        f"{label_p}最高收盘价, 日期: {stats.get(max_date_key, '未知')}")
            if crit_key in stats:
                add_row(f"建议临界买点（{label_p}）", f"{stats[crit_key]:.2f} 元",
                        f"{label_p}最小值×{stats.get('adjusted_premium', 1.15):.2f}, 溢价{stats.get(prem_key, 0):.1f}%")
    else:
        add_row(f"价格最小值（近{years}年）", f"{stats['min']:.2f} 元",
                f"统计周期内最低收盘价, 日期: {stats.get('min_date_str', '未知')}")
        add_row(f"价格最大值（近{years}年）", f"{stats['max']:.2f} 元",
                f"统计周期内最高收盘价, 日期: {stats.get('max_date_str', '未知')}")
        crit = stats.get("critical_buy_point", 0)
        prem = stats.get("buy_point_premium", 0)
        add_row("建议临界买点", f"{crit:.2f} 元",
                f"最小值×1.15, 相比最低价溢价{prem:.1f}%")

    latest_date = stats.get("latest_trade_date", "未知")
    add_row(f"最新收盘价（{latest_date}）", f"{stats.get('latest_price', 0):.2f} 元",
            "统计周期内最后一个交易日的收盘价")
    add_row("最新价与均值绝对差异", f"{stats.get('price_diff_abs', 0):.2f} 元",
            "最新价 - 均值，正数代表最新价高于均值")
    add_row("最新价与均值相对差异", f"{stats.get('price_diff_pct', 0):.2f}%",
            "(最新价-均值)/均值，反映最新价相对均值的偏离程度")
    add_row("均值±1倍标准差区间",
            f"[{stats['std1_lower']:.2f}, {stats['std1_upper']:.2f}] 元",
            "约68%数据在此区间")
    add_row("均值±2倍标准差区间",
            f"[{stats['std2_lower']:.2f}, {stats['std2_upper']:.2f}] 元",
            "约95%数据在此区间")

    return pd.DataFrame(rows, columns=["统计指标", "数值/区间", "统计学说明"])


def build_investment_advice_sheet(stats: dict, multi_period: bool = False) -> pd.DataFrame:
    """生成「投资建议」Sheet 的 DataFrame。"""
    rows: list[list] = []
    latest = stats.get("latest_price", 0)

    rows.append(["投资建议", ""])
    rows.append(["最新价", f"{latest:.2f} 元"])

    if multi_period:
        for period, label in [("1y", "近1年"), ("2y", "近2年"), ("3y", "近3年")]:
            crit_key = f"critical_buy_point_{period}"
            if crit_key in stats:
                rows.append([f"{label}临界点", f"{stats[crit_key]:.2f} 元"])
        rows.append(["", ""])
        rows.append(["位置判断", ""])
        for period, label in [("1y", "近1年"), ("2y", "近2年"), ("3y", "近3年")]:
            crit_key = f"critical_buy_point_{period}"
            if crit_key in stats:
                pos = "低于" if latest <= stats[crit_key] else "高于"
                rows.append([f"相对{label}临界点", pos])
    else:
        crit = stats.get("critical_buy_point", 0)
        rows.append(["临界买点", f"{crit:.2f} 元"])
        rows.append(["", ""])
        rows.append(["位置判断", ""])
        pos = "低于" if latest <= crit else "高于"
        rows.append(["相对临界点", pos])
        rows.append(["价格差异", f"{abs(latest - crit):.2f} 元"])

    rows.append(["", ""])
    rows.append(["投资建议", ""])
    return pd.DataFrame(rows, columns=["项目", "数值"])
